#!/usr/bin/env python3
"""
DCF model validator.

Opens an Excel DCF workbook and checks it for spreadsheet formula errors
and a handful of DCF sanity rules (terminal growth vs. WACC, WACC range,
terminal-value share of enterprise value). Emits a JSON report.
"""

import sys
import json
from datetime import datetime
from pathlib import Path


# Excel cell error tokens we scan every cell for.
EXCEL_ERROR_TOKENS = ("#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A")


class DCFModelValidator:
    """Runs structural and DCF-logic checks against a workbook."""

    def __init__(self, excel_path):
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl")

        if not Path(excel_path).exists():
            raise FileNotFoundError(f"File not found: {excel_path}")

        self.excel_path = excel_path
        self.openpyxl = openpyxl
        # One handle keeps the raw formulas, the other the last computed values.
        self.workbook_formulas = openpyxl.load_workbook(excel_path, data_only=False)
        self.workbook_values = openpyxl.load_workbook(excel_path, data_only=True)

        self.errors = []
        self.warnings = []
        self.info = []

    def validate_all(self):
        """Run every check and return the assembled report dict."""
        self.check_sheet_structure()
        self.check_formula_errors()
        self.check_dcf_logic()

        return {
            "file": self.excel_path,
            "validation_date": datetime.now().isoformat(),
            "status": "PASS" if not self.errors else "FAIL",
            "error_count": len(self.errors),
            "warning_count": len(self.warnings),
            "errors": self.errors,
            "warnings": self.warnings,
            "info": self.info,
        }

    # ------------------------------------------------------------------ #
    # structural checks
    # ------------------------------------------------------------------ #

    def check_sheet_structure(self):
        """Note which of the recommended sheets are present."""
        present = set(self.workbook_values.sheetnames)
        for sheet in ("DCF", "WACC", "Sensitivity"):
            if sheet in present:
                self.info.append(f"Found sheet: {sheet}")
            else:
                self.warnings.append(f"Recommended sheet missing: {sheet}")

    def check_formula_errors(self):
        """Walk every cell of every sheet, counting formulas and error tokens."""
        locations_by_error = {token: [] for token in EXCEL_ERROR_TOKENS}
        total_errors = 0
        total_formulas = 0

        for sheet_name in self.workbook_values.sheetnames:
            value_sheet = self.workbook_values[sheet_name]
            formula_sheet = self.workbook_formulas[sheet_name]

            for row in value_sheet.iter_rows():
                for cell in row:
                    raw = formula_sheet[cell.coordinate].value
                    if isinstance(raw, str) and raw.startswith("="):
                        total_formulas += 1

                    computed = cell.value
                    if isinstance(computed, str):
                        for token in EXCEL_ERROR_TOKENS:
                            if token in computed:
                                where = f"{sheet_name}!{cell.coordinate}"
                                locations_by_error[token].append(where)
                                total_errors += 1
                                self.errors.append(f"{token} at {where}")
                                break

        self.info.append(f"Total formulas: {total_formulas}")
        if total_errors == 0:
            self.info.append("✓ No formula errors found")
        else:
            self.errors.append(f"Total formula errors: {total_errors}")

        return locations_by_error, total_errors

    # ------------------------------------------------------------------ #
    # DCF-specific logic checks
    # ------------------------------------------------------------------ #

    def check_dcf_logic(self):
        """Run the three DCF reasonableness checks."""
        self._check_terminal_growth_vs_wacc()
        self._check_wacc_range()
        self._check_terminal_value_proportion()

    @staticmethod
    def _first_fraction_to_the_right(sheet, cell, span=4):
        """
        Return the first numeric value strictly between 0 and 1 found in the
        cells immediately to the right of `cell` (within `span` columns), or None.
        """
        for offset in range(1, span + 1):
            value = sheet.cell(cell.row, cell.column + offset).value
            if isinstance(value, (int, float)) and 0 < value < 1:
                return value
        return None

    @staticmethod
    def _first_positive_to_the_right(sheet, cell, span=4):
        """First strictly-positive numeric value to the right of `cell`, or None."""
        for offset in range(1, span + 1):
            value = sheet.cell(cell.row, cell.column + offset).value
            if isinstance(value, (int, float)) and value > 0:
                return value
        return None

    def _check_terminal_growth_vs_wacc(self):
        """The hard constraint: terminal growth must be below WACC."""
        try:
            dcf = self.workbook_values["DCF"]
        except KeyError:
            self.warnings.append("DCF sheet not found")
            return

        try:
            terminal_growth = None
            wacc = None

            for row in dcf.iter_rows(max_row=100, max_col=20):
                for cell in row:
                    if not isinstance(cell.value, str):
                        continue
                    label = cell.value.lower()
                    if "terminal" in label and "growth" in label and terminal_growth is None:
                        terminal_growth = self._first_fraction_to_the_right(dcf, cell)
                    if "wacc" in label and wacc is None:
                        wacc = self._first_fraction_to_the_right(dcf, cell)

            if terminal_growth is None or wacc is None:
                self.warnings.append("Could not locate terminal growth and WACC values")
            elif terminal_growth >= wacc:
                self.errors.append(
                    f"CRITICAL: Terminal growth ({terminal_growth:.2%}) >= WACC ({wacc:.2%}). "
                    "This creates infinite value and is mathematically invalid."
                )
            else:
                self.info.append(
                    f"✓ Terminal growth ({terminal_growth:.2%}) < WACC ({wacc:.2%})"
                )
        except Exception as exc:
            self.warnings.append(f"Could not validate terminal growth vs WACC: {exc}")

    def _check_wacc_range(self):
        """WACC outside roughly 5%-20% is worth a second look."""
        try:
            sheet = self.workbook_values["WACC"] if "WACC" in self.workbook_values.sheetnames \
                else self.workbook_values["DCF"]

            wacc = None
            for row in sheet.iter_rows(max_row=100, max_col=20):
                for cell in row:
                    if isinstance(cell.value, str) and "wacc" in cell.value.lower():
                        found = self._first_fraction_to_the_right(sheet, cell)
                        if found is not None:
                            wacc = found

            if wacc is None:
                self.warnings.append("Could not locate WACC value")
            elif wacc < 0.05 or wacc > 0.20:
                self.warnings.append(
                    f"WACC ({wacc:.2%}) is outside typical range (5%-20%). Verify calculation."
                )
            else:
                self.info.append(f"✓ WACC ({wacc:.2%}) in reasonable range")
        except Exception as exc:
            self.warnings.append(f"Could not validate WACC range: {exc}")

    def _check_terminal_value_proportion(self):
        """Terminal value should land at roughly 40%-80% of enterprise value."""
        try:
            dcf = self.workbook_values["DCF"]

            terminal_value = None
            enterprise_value = None

            for row in dcf.iter_rows(max_row=200, max_col=20):
                for cell in row:
                    if not isinstance(cell.value, str):
                        continue
                    label = cell.value.lower()
                    if "terminal" in label and "value" in label and "pv" in label \
                            and terminal_value is None:
                        terminal_value = self._first_positive_to_the_right(dcf, cell)
                    if "enterprise" in label and "value" in label and enterprise_value is None:
                        enterprise_value = self._first_positive_to_the_right(dcf, cell)

            if terminal_value is None or not enterprise_value:
                self.warnings.append("Could not locate terminal value and enterprise value")
                return

            share = terminal_value / enterprise_value
            if share > 0.80:
                self.warnings.append(
                    f"Terminal value is {share:.1%} of EV (typically should be 50-70%). "
                    "Model may be over-reliant on terminal assumptions."
                )
            elif share < 0.40:
                self.warnings.append(
                    f"Terminal value is {share:.1%} of EV (typically should be 50-70%). "
                    "Check if terminal assumptions are too conservative."
                )
            else:
                self.info.append(f"✓ Terminal value is {share:.1%} of EV")
        except Exception as exc:
            self.warnings.append(f"Could not validate terminal value proportion: {exc}")


def validate_dcf_model(excel_path):
    """Convenience wrapper: build a validator and return its report."""
    return DCFModelValidator(excel_path).validate_all()


def main():
    """Command-line entry point."""
    if len(sys.argv) < 2:
        print("Usage: python validate_dcf.py <excel_file> [output.json]")
        print("\nValidates DCF model for:")
        print("  - Formula errors (#REF!, #DIV/0!, etc.)")
        print("  - Terminal growth < WACC (critical)")
        print("  - WACC in reasonable range (5-20%)")
        print("  - Terminal value proportion of EV (40-80%)")
        print("\nReturns JSON with errors, warnings, and info")
        print("\nExample: python validate_dcf.py model.xlsx")
        print("Example: python validate_dcf.py model.xlsx results.json")
        sys.exit(1)

    excel_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None

    try:
        results = validate_dcf_model(excel_file)
        print(json.dumps(results, indent=2))

        if output_file:
            with open(output_file, "w") as handle:
                json.dump(results, handle, indent=2)

        sys.exit(0 if results["status"] == "PASS" else 1)
    except Exception as exc:
        print(json.dumps({"file": excel_file, "status": "ERROR", "error": str(exc)}, indent=2))
        sys.exit(1)


if __name__ == "__main__":
    main()
